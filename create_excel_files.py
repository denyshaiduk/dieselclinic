import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Data definition matches the prompt exactly
data = {
    "ОД-52392": {
        "buyer_supplier": "Покупець: бар Бункер",
        "vat": 371.52,
        "items": [
            ("Напій безалкогольний \"Кока-кола\" 0,33л ж/б 12шт", "5449000000996", 3, 36, 26.01, 936.36),
            ("Ананаси різані ж/б ТМ Чемпіон 850мл", "482014948099", 0.04, 1, 100.01, 100.01),
            ("Ананасі кільця ж/б ТМ Чемпіон 850мл 24шт", "4820068980733", 0.04, 1, 106.78, 106.78),
            ("Напій безалкогольний \"Швепс\" Мохіто 1л", "5449000109125", 0.33, 4, 43.10, 172.40),
            ("Напій безалкогольний \"Швепс Тонік\" 1л", "5449000044808", 0.33, 4, 43.10, 172.40),
            ("Напій безалкогольний \"Швепс Бітер Лимон\" 1л", "5449000044839", 0.33, 4, 43.10, 172.40),
            ("Напій безалкогольний \"Спрайт\" 1,25л", "5449000028976", 1, 6, 46.41, 278.46),
            ("Напій безалкогольний \"Кока-кола\" 1,25л", "5449000028921", 1, 6, 48.39, 290.34)
        ]
    },
    "26653": {
        "buyer_supplier": "постачальник ТОВ ТТ",
        "vat": 2498.76,
        "items": [
            ("Горілка ФІНЛЯНДІЯ 1л", "6412700021027", 0.4, 5, 593.69, 2968.45),
            ("Віскі бурбон \"Джим Бім Вайт\" 40% 1л", "5010196092142", 0.25, 3, 1057.42, 3172.26),
            ("Виски ДЖЕЙМСОН 1л", "5011007003227", 0.5, 3, 964.74, 2894.22),
            ("Віскі \"Джек Деніелс\" 1л", "5099873045367", 0.25, 3, 1164.55, 3493.65),
            ("Горілка Немиров Особливий 1л штоф укр", "4820181420635", 0.25, 3, 253.56, 760.68),
            ("Горілка Немиров Делікат м'яка 1л штоф", "4820181420451", 0.25, 3, 253.56, 760.68),
            ("Вино ігристе Martini Asti DOCG 7.5% 0.75л", "8000570435402", 0.17, 1, 538.02, 538.02),
            ("Вермут Мартіні Бианко 1л", "5010677925006", 0.17, 1, 404.58, 404.58)
        ]
    },
    "26654": {
        "buyer_supplier": "постачальник ТОВ ТТ",
        "vat": 163.47,
        "items": [
            ("Пиво КАРЛСБЕРГ 0.45л", "4820000458795", 1, 20, 49.04, 980.80)
        ]
    },
    "26652": {
        "buyer_supplier": "постачальник ТОВ ТТ",
        "vat": 1900.13,
        "items": [
            ("Горілка Козацька Рада Особлива 1л", "4820080724087", 0.4, 5, 196.96, 984.80),
            ("Алкогольний напій Mertes Sparkling газ н/сол білий 0.75л", "4003301071881", 0.3, 2, 228.53, 457.06),
            ("Алкогольний напій Peter Mertes Sparkling Breeze газ н/сух білий 0.75л", "4003301081552", 0.3, 2, 220.27, 440.54),
            ("Алкогольний напій Zubrowka Bison Grass 37.5% 0.7л", "5900343003698", 0.25, 3, 192.48, 577.44),
            ("Горілка Zubrowka Czarna 40% 0.7л", "5900343010054", 0.25, 3, 316.92, 950.76),
            ("Горілка Zubrowka Biala 40% 1л", "5900343001939", 0.17, 2, 258.18, 516.36),
            ("Віскі White Horse 1л", "5000265001335", 0.08, 1, 651.95, 651.95),
            ("Лікер Sheridans 0.7л", "5011013500680", 0.33, 2, 1000.31, 2000.62),
            ("Лікер Jagermeister 1л", "4067700013002", 0.67, 4, 810.39, 3241.56),
            ("Джин Gordon's (37.5%) 1л", "5000289020800", 0.17, 2, 789.83, 1579.66)
        ]
    }
}

# Verification
print("--- STARTING DATA VERIFICATION ---")
discrepancies = []

for sheet_name, sheet_data in data.items():
    print(f"\nVerifying delivery note: {sheet_name} ({sheet_data['buyer_supplier']})")
    items = sheet_data["items"]
    sheet_sum = 0.0
    for idx, item in enumerate(items, 1):
        name, barcode, qty_in_pack, qty, price, total = item
        calc_total = round(qty * price, 2)
        if abs(calc_total - total) > 0.01:
            err = f"  Mismatch at item {idx} ({name}): Qty ({qty}) * Price ({price}) = {calc_total}, but given total is {total}."
            print(f"  [ERROR] {err}")
            discrepancies.append(err)
        sheet_sum += total
    
    sheet_sum = round(sheet_sum, 2)
    expected_sum = {
        "ОД-52392": 2229.15,
        "26653": 14992.54,
        "26654": 980.80,
        "26652": 11400.75
    }[sheet_name]
    
    if abs(sheet_sum - expected_sum) > 0.01:
        err = f"  Mismatch in sheet {sheet_name} total sum: Sum of items = {sheet_sum}, but expected sheets total = {expected_sum}."
        print(f"  [ERROR] {err}")
        discrepancies.append(err)
    else:
        print(f"  Success: Sheet grand total ({sheet_sum}) matches sum of all rows perfectly!")

if not discrepancies:
    print("\n[SUCCESS] Verification complete. All row totals and sheet grand totals match perfectly! NO mismatches found.\n")
else:
    print(f"\n[WARNING] Verification complete. Found {len(discrepancies)} mismatches.\n")


# Set up reusable styles
font_family = "Calibri"
font_regular = Font(name=font_family, size=11, bold=False)
font_bold = Font(name=font_family, size=11, bold=True)
font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")

fill_header = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")

thin_side = Side(border_style="thin", color="D3D3D3")
border_data = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

format_integer = "#,##0"
format_decimal = "#,##0.00"

# --- GENERATE FILE 1: накладні_07.08.2026.xlsx ---
print("Generating FILE 1: накладні_07.08.2026.xlsx")
wb1 = openpyxl.Workbook()
wb1.remove(wb1.active)

for sheet_name, sheet_data in data.items():
    ws = wb1.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = True
    
    # Freeze row 2 (so that row 1 containing headers is frozen)
    ws.freeze_panes = "A2"
    
    # Write Header row
    headers = ["№", "Товар", "Штрихкод", "К-ть в уп.", "Кількість", "Ціна з ПДВ", "Сума з ПДВ"]
    ws.row_dimensions[1].height = 24
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx not in [2] else align_left
        cell.border = border_data
    
    # Write Data rows
    items = sheet_data["items"]
    for row_idx, item in enumerate(items, 2):
        name, barcode, qty_in_pack, qty, price, total = item
        ws.row_dimensions[row_idx].height = 20
        
        # 1. №
        c1 = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        c1.font = font_regular
        c1.alignment = align_center
        c1.number_format = format_integer
        c1.border = border_data
        
        # 2. Товар
        c2 = ws.cell(row=row_idx, column=2, value=name)
        c2.font = font_regular
        c2.alignment = align_left
        c2.border = border_data
        
        # 3. Штрихкод
        c3 = ws.cell(row=row_idx, column=3, value=str(barcode))
        c3.font = font_regular
        c3.alignment = align_center
        c3.number_format = "@"
        c3.border = border_data
        
        # 4. К-ть в уп.
        c4 = ws.cell(row=row_idx, column=4, value=qty_in_pack)
        c4.font = font_regular
        c4.alignment = align_right
        c4.number_format = "0.00"
        c4.border = border_data
        
        # 5. Кількість
        c5 = ws.cell(row=row_idx, column=5, value=qty)
        c5.font = font_regular
        c5.alignment = align_right
        c5.number_format = format_integer if int(qty) == qty else "0.00"
        c5.border = border_data
        
        # 6. Ціна з ПДВ
        c6 = ws.cell(row=row_idx, column=6, value=price)
        c6.font = font_regular
        c6.alignment = align_right
        c6.number_format = format_decimal
        c6.border = border_data
        
        # 7. Сума з ПДВ (calculated)
        c7 = ws.cell(row=row_idx, column=7, value=total)
        c7.font = font_regular
        c7.alignment = align_right
        c7.number_format = format_decimal
        c7.border = border_data
        
    last_item_row = len(items) + 1
    totals_row_idx = last_item_row + 1
    vat_row_idx = totals_row_idx + 1
    
    # Totals Row (Разом)
    ws.row_dimensions[totals_row_idx].height = 20
    for col_idx in range(1, 8):
        c = ws.cell(row=totals_row_idx, column=col_idx)
        c.border = border_data
        
    lbl_c = ws.cell(row=totals_row_idx, column=2, value="Разом:")
    lbl_c.font = font_bold
    lbl_c.alignment = align_left
    
    sum_c = ws.cell(row=totals_row_idx, column=7, value=f"=SUM(G2:G{last_item_row})")
    sum_c.font = font_bold
    sum_c.alignment = align_right
    sum_c.number_format = format_decimal
    
    # VAT Row (У т.ч. ПДВ)
    ws.row_dimensions[vat_row_idx].height = 20
    for col_idx in range(1, 8):
        c = ws.cell(row=vat_row_idx, column=col_idx)
        c.border = border_data
        
    vat_lbl = ws.cell(row=vat_row_idx, column=2, value="У т.ч. ПДВ:")
    vat_lbl.font = font_bold
    vat_lbl.alignment = align_left
    
    vat_val = ws.cell(row=vat_row_idx, column=7, value=sheet_data["vat"])
    vat_val.font = font_bold
    vat_val.alignment = align_right
    vat_val.number_format = format_decimal
    
    # Set column widths
    col_widths = {
        1: 6,   # №
        2: 45,  # Товар
        3: 18,  # Штрихкод
        4: 12,  # К-ть в уп.
        5: 12,  # Кількість
        6: 15,  # Ціна з ПДВ
        7: 15   # Сума з ПДВ
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

wb1.save("накладні_07.08.2026.xlsx")
print("Saved FILE 1 successfully.")


# --- GENERATE FILE 2: інвентаризація_бар_07.08.2026.xlsx ---
print("\nGenerating FILE 2: інвентаризація_бар_07.08.2026.xlsx")
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Інвентаризація"
ws2.sheet_view.showGridLines = True

# Freeze row 2
ws2.freeze_panes = "A2"

# Headers for File 2
headers2 = [
    "№", "Накладна", "Штрихкод", "Товар", 
    "Кількість за накладною", "Ціна з ПДВ", "Сума з ПДВ", 
    "Фактична кількість", "Статус", "Коментар"
]
ws2.row_dimensions[1].height = 24
for col_idx, header_text in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header_text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center if col_idx not in [4, 10] else align_left
    cell.border = border_data

# Combine all 27 items from all 4 sheets in order
row_idx = 2
for sheet_name, sheet_data in data.items():
    items = sheet_data["items"]
    for item in items:
        name, barcode, qty_in_pack, qty, price, total = item
        ws2.row_dimensions[row_idx].height = 20
        
        # A: № (1-27)
        c_no = ws2.cell(row=row_idx, column=1, value=row_idx - 1)
        c_no.font = font_regular
        c_no.alignment = align_center
        c_no.number_format = format_integer
        c_no.border = border_data
        
        # B: Накладна
        c_inv = ws2.cell(row=row_idx, column=2, value=sheet_name)
        c_inv.font = font_regular
        c_inv.alignment = align_center
        c_inv.number_format = "@"
        c_inv.border = border_data
        
        # C: Штрихкод
        c_bc = ws2.cell(row=row_idx, column=3, value=str(barcode))
        c_bc.font = font_regular
        c_bc.alignment = align_center
        c_bc.number_format = "@"
        c_bc.border = border_data
        
        # D: Товар
        c_name = ws2.cell(row=row_idx, column=4, value=name)
        c_name.font = font_regular
        c_name.alignment = align_left
        c_name.border = border_data
        
        # E: Кількість за накладною
        c_qty = ws2.cell(row=row_idx, column=5, value=qty)
        c_qty.font = font_regular
        c_qty.alignment = align_right
        c_qty.number_format = format_integer if int(qty) == qty else "0.00"
        c_qty.border = border_data
        
        # F: Ціна з ПДВ
        c_pr = ws2.cell(row=row_idx, column=6, value=price)
        c_pr.font = font_regular
        c_pr.alignment = align_right
        c_pr.number_format = format_decimal
        c_pr.border = border_data
        
        # G: Сума з ПДВ
        c_tot = ws2.cell(row=row_idx, column=7, value=total)
        c_tot.font = font_regular
        c_tot.alignment = align_right
        c_tot.number_format = format_decimal
        c_tot.border = border_data
        
        # H: Фактична кількість (empty)
        c_fq = ws2.cell(row=row_idx, column=8, value=None)
        c_fq.font = font_regular
        c_fq.alignment = align_right
        c_fq.border = border_data
        
        # I: Статус (empty cell with dropdown)
        c_st = ws2.cell(row=row_idx, column=9, value=None)
        c_st.font = font_regular
        c_st.alignment = align_center
        c_st.border = border_data
        
        # J: Коментар (empty)
        c_cmt = ws2.cell(row=row_idx, column=10, value=None)
        c_cmt.font = font_regular
        c_cmt.alignment = align_left
        c_cmt.border = border_data
        
        row_idx += 1

# Add data validation to column I (Статус), rows 2-28
dv = DataValidation(
    type="list", 
    formula1='"В наявності,Немає акцизу,Пляшка закінчилась,Нестача,Лишок,Пересорт"', 
    allow_blank=True
)
dv.error = "Оберіть значення зі списку"
dv.errorTitle = "Помилка введення"
dv.prompt = "Оберіть статус"
dv.promptTitle = "Статус"

ws2.add_data_validation(dv)
dv.add("I2:I28")

# Totals Row after row 28 (i.e. row 29)
totals_row_idx = 29
ws2.row_dimensions[totals_row_idx].height = 20
for col_idx in range(1, 11):
    c = ws2.cell(row=totals_row_idx, column=col_idx)
    c.border = border_data

# Bold "Разом:" in column F
lbl_f = ws2.cell(row=totals_row_idx, column=6, value="Разом:")
lbl_f.font = font_bold
lbl_f.alignment = align_right

# formula =SUM(G2:G28) in column G
sum_g = ws2.cell(row=totals_row_idx, column=7, value="=SUM(G2:G28)")
sum_g.font = font_bold
sum_g.alignment = align_right
sum_g.number_format = format_decimal

# Column widths
col_widths2 = {
    1: 6,    # №
    2: 12,   # Накладна
    3: 18,   # Штрихкод
    4: 45,   # Товар
    5: 22,   # Кількість за накладною
    6: 15,   # Ціна з ПДВ
    7: 15,   # Сума з ПДВ
    8: 20,   # Фактична кількість
    9: 20,   # Статус
    10: 25   # Коментар
}

for col_idx, width in col_widths2.items():
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

wb2.save("інвентаризація_бар_07.08.2026.xlsx")
print("Saved FILE 2 successfully.")
